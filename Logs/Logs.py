import time
import openpyxl


class Report:

    def log(self, message):
        current_time = time.strftime('%H:%M:%S')
        print(f'{current_time}  {message}')

    def append_report(self, data, report_path):
        rep = openpyxl.load_workbook(report_path)
        sheet = rep['Report']
        row = 3
        while sheet.cell(row=row, column=1).value:
            row += 1
        sheet.cell(row=row, column=1, value=data['audio'])
        sheet.cell(row=row, column=2, value=data['wer'])
        sheet.cell(row=row, column=3, value=data['capitalization']['support'])
        sheet.cell(row=row, column=4, value=data['capitalization']['precision'])
        sheet.cell(row=row, column=5, value=data['capitalization']['recall'])
        sheet.cell(row=row, column=6, value=data['capitalization']['f1'])
        sheet.cell(row=row, column=7, value=data['no_capitalization']['support'])
        sheet.cell(row=row, column=8, value=data['no_capitalization']['precision'])
        sheet.cell(row=row, column=9, value=data['no_capitalization']['recall'])
        sheet.cell(row=row, column=10, value=data['no_capitalization']['f1'])
        sheet.cell(row=row, column=11, value=data['punctuation']['support'])
        sheet.cell(row=row, column=12, value=data['punctuation']['precision'])
        sheet.cell(row=row, column=13, value=data['punctuation']['recall'])
        sheet.cell(row=row, column=14, value=data['punctuation']['f1'])
        sheet.cell(row=row, column=15, value=data['no_punctuation']['support'])
        sheet.cell(row=row, column=16, value=data['no_punctuation']['precision'])
        sheet.cell(row=row, column=17, value=data['no_punctuation']['recall'])
        sheet.cell(row=row, column=18, value=data['no_punctuation']['f1'])
        sheet.cell(row=row, column=19, value=data['period']['support'])
        sheet.cell(row=row, column=20, value=data['period']['precision'])
        sheet.cell(row=row, column=21, value=data['period']['recall'])
        sheet.cell(row=row, column=22, value=data['period']['f1'])
        sheet.cell(row=row, column=23, value=data['comma']['support'])
        sheet.cell(row=row, column=24, value=data['comma']['precision'])
        sheet.cell(row=row, column=25, value=data['comma']['recall'])
        sheet.cell(row=row, column=26, value=data['comma']['f1'])
        sheet.cell(row=row, column=27, value=data['question']['support'])
        sheet.cell(row=row, column=28, value=data['question']['precision'])
        sheet.cell(row=row, column=29, value=data['question']['recall'])
        sheet.cell(row=row, column=30, value=data['question']['f1'])
        sheet.cell(row=row, column=31, value=data['exclamation']['support'])
        sheet.cell(row=row, column=32, value=data['exclamation']['precision'])
        sheet.cell(row=row, column=33, value=data['exclamation']['recall'])
        sheet.cell(row=row, column=34, value=data['exclamation']['f1'])
        sheet.cell(row=row, column=35, value=data['column']['support'])
        sheet.cell(row=row, column=36, value=data['column']['precision'])
        sheet.cell(row=row, column=37, value=data['column']['recall'])
        sheet.cell(row=row, column=38, value=data['column']['f1'])
        sheet.cell(row=row, column=39, value=data['ellipsis']['support'])
        sheet.cell(row=row, column=40, value=data['ellipsis']['precision'])
        sheet.cell(row=row, column=41, value=data['ellipsis']['recall'])
        sheet.cell(row=row, column=42, value=data['ellipsis']['f1'])
        sheet.cell(row=row, column=43, value=data['stt']['expected'])
        sheet.cell(row=row, column=44, value=data['stt']['actual'])
        sheet.cell(row=row, column=45, value=data['nlp']['expected'])
        sheet.cell(row=row, column=46, value=data['nlp']['actual'])
        rep.save(report_path)
        rep.close()
